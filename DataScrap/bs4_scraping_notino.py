from bs4 import BeautifulSoup
import requests
import pathlib
import openpyxl
import re
import os
import numpy as np
import pandas as pd
main_link_notino="https://www.notino.fr/"
page=requests.get(main_link_notino)
soup=BeautifulSoup(page.content,'html.parser')
data=pd.read_excel('data.xlsx')
def chgetInfo(product):
    return product.columns
def chget(val,product):   
    return {_:product[_][val] for _ in product.columns}
def makeRef(brand,ref1,ref2):
    def globchange(x):
        x=x.replace("’","")
        x=x.replace("é","e")
        x=x.replace("è","e")
        x=x.replace("ê","e")
        x=x.replace("ô","o")
        x=x.replace("!","")
        x=x.replace("(","")
        x=x.replace(")","")
        return x
    ref1=globchange(ref1)
    brand=globchange(brand)
    ref2=globchange(ref2)
    return {brand:{ref1:ref2}}
def diff_brand_and_ref(a,b):
    return b.replace(a+" ","")
class bs4_scraping_notino:
        def __init__(self):
            pass
        
        def Xtract_list_OFLinkOF_man_perfume(self):
            link_man_perfum="https://www.notino.fr/parfums-homme/"
            page=requests.get(link_man_perfum)
            dicte={}
            k=24
            for j in range(500*24):
                if j%k==0:
                    link_man_perfum="https://www.notino.fr/parfums-homme/?f="+str(j//k)+"-1-55544-55549"
                    page=requests.get(link_man_perfum)
                    soup=BeautifulSoup(page.content,'html.parser')
                    print(j)
                    k=len(soup)
                try:
                    sp=soup.find_all('h2',class_="sc-gWXbKe sc-dvQaRk jUthAs jJrJMY")
                    brand=sp[j%k].text              
                    sl=soup.find_all('h3',class_="sc-cCcXHH sc-TBWPX iRiafX dAontC")
                    ref1=sl[j%k].text 
                    sa=soup.find_all('p',class_="sc-ZOtfp etDeHy")
                    ref2=sa[j%k].text 
                    res=makeRef(brand,ref1,ref2)
                    dicte[j]=res
                except:
                    print("page"+str(j//k)+"sauté")
                    
       
            return dicte
 
        def Xtract_list_OFLinkOF_woman_perfume(self):
            link_women_perfum="https://www.notino.fr/parfums-femme/"
            page=requests.get(link_women_perfum)
            soup=BeautifulSoup(page.content,'html.parser')
            dicte={}
            k=24
            for j in range(500*24):
                if j%k==0:
                    link_man_perfum="https://www.notino.fr/parfums-femme/?f="+str((j//k)%96)+"-1-55544-55545"
                    page=requests.get(link_man_perfum)
                    soup=BeautifulSoup(page.content,'html.parser')
                    print(j)
                    k=len(soup)
                try:
                    sp=soup.find_all('h2',class_="sc-gWXbKe sc-dvQaRk jUthAs jJrJMY")
                    brand=sp[j%k].text              
                    sl=soup.find_all('h3',class_="sc-cCcXHH sc-TBWPX iRiafX dAontC")
                    ref1=sl[j%k].text 
                    sa=soup.find_all('p',class_="sc-ZOtfp etDeHy")
                    ref2=sa[j%k].text 
                    res=makeRef(brand,ref1,ref2)
                    dicte[j]=res
                except:
                    print("page"+str(j//k)+"sauté")
                    
       
            return dicte
        def testctrat(self):
            link_women_perfum="https://www.notino.fr/parfums-femme/"
            page=requests.get(link_women_perfum)
            soup=BeautifulSoup(page.content,'html.parser')
            dicte={}
            k=24
            for j in range(24):
                if j%k==0:
                    link_man_perfum="https://www.notino.fr/parfums-femme/?f="+str((j//k)%139)+"-1-55544-55545"
                    page=requests.get(link_man_perfum)
                    soup=BeautifulSoup(page.content,'html.parser')
                    print(j)
                    k=len(soup)
                try:
                    sp=soup.find_all('h2',class_="sc-gWXbKe sc-dvQaRk jUthAs jJrJMY")
                    brand=sp[j%k].text              
                    sl=soup.find_all('h3',class_="sc-cCcXHH sc-TBWPX iRiafX dAontC")
                    ref1=sl[j%k].text 
                    sa=soup.find_all('p',class_="sc-ZOtfp etDeHy")
                    ref2=sa[j%k].text 
                    res=makeRef(brand,ref1,ref2)
                    dicte[j]=res
                except:
                    print("page"+str(j//k)+"sauté")
                    
       
            return dicte
        def proc_all_link(self,dict1,dict2):
            liste_of_link=[]
            for i in dict1.values():
                for c,b in i.items():
                    for j in b:
                        link=c+"/"+j+" "+b[j]
                        link=link.replace(" ","-")
                liste_of_link.append(link)
            for i in dict2.values():
                for c,b in i.items():
                    for j in b:
                        link=c+"/"+j+" "+b[j]
                        link=link.replace(" ","-")
                liste_of_link.append(link)
            return ["https://www.notino.fr/"+i for i in liste_of_link]                
        def get_ingredients_of_perfume(self,link):
            
            page=requests.get(link)
            soup=BeautifulSoup(page.content,'html.parser')
            sp=soup.find_all("div",class_="styled__Description-sc-1eu1dd2-3 styled__ExpandableDescription-sc-1f6wyem-0 eZoNhS ghGzLS")
            print(sp)
            for i in sp:
                k=i.text.split(",")
                if len(k)<5:
                    k=k=i.text.split(".")
                ing=k[:-3]
                if ing is None:
                    return []
                else:
                    return ing
                
        def get_rating(self,link):
            
            page=requests.get(link)
            soup=BeautifulSoup(page.content,'html.parser')
            sp=soup.find_all("a",class_="styled__ReviewsAnchor-sc-3sotvb-7 dla-DMn")
            for i in sp:
                str_sp=str(sp)
                index_sp=str_sp.find("title=")
                if str_sp[index_sp+8]==".":
                    rating=float(str_sp[index_sp+7:82])
                else :
                    rating=float(str_sp[index_sp+7])
                print(rating)
                return rating
       
def makeLinkNotino():
    path =str(pathlib.Path(__file__).parent.absolute())+'\data.xlsx'
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    notino_data=bs4_scraping_notino()
    dict_man_perfume=notino_data.Xtract_list_OFLinkOF_man_perfume()

    dict_women_perfume=notino_data.Xtract_list_OFLinkOF_woman_perfume()
    liste_of_link=notino_data.proc_all_link(dict_women_perfume,dict_man_perfume)
    liste_of_link=set(liste_of_link)
    for e,i in enumerate(liste_of_link):
        lk=sheet_obj.cell(row=e+2,column=1)
        lk.value=i
    wb_obj.save('data.xlsx')
def makeNotino():
     
    for i in range(chgetInfo(data)):
        
        ing=notino_data.get_ingredients_of_perfume(i)
        print(ing)
        rat=notino_data.get_rating(i)
        dicte[i]=[ing,rat]
    return dicte,len(liste_of_link)
def makeIngredient(rng):
    path =str(pathlib.Path(__file__).parent.absolute())+'\data.xlsx'
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    scrap=bs4_scraping_notino()
    for i in rng:
        ingr=scrap.get_ingredients_of_perfume(chget(i,data)["Name"])
        if pd.isnull(chget(i,data)["Ingr"]):
            lk=sheet_obj.cell(row=i+2,column=2)
            if ingr!=None:
                lk.value=",".join(ingr)
            else:
                pass
        else:
            pass
    wb_obj.save('data.xlsx')
def makeRating(rng):
    path =str(pathlib.Path(__file__).parent.absolute())+'\data.xlsx'
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    scrap=bs4_scraping_notino()
    for i in rng:
        ingr=scrap.get_rating(chget(i,data)["Name"])
        lk=sheet_obj.cell(row=i+2,column=3)
        if np.isnan(chget(i,data)["Rating"]):
            
            if ingr!=None:
                lk.value=ingr
                
            else:
                pass
    wb_obj.save('data.xlsx')
  #  for e,i in enumerate(l):
   #     lk=sheet_obj.cell(row=e+2,column=1)
    #    lk.value=i
   # wb_obj.save('data.xlsx')
def makeReference(link):
    ref=link.split("/")[-1]
    return ref.replace("-"," ")
def makeBrand(link):
    brand=link.split("/")[3]
    return brand.replace("-"," ")
        
#data=bs4_scraping_notino()
#a=data.get_ingredients_of_perfume("https://www.notino.fr/calvin-klein/ck-be-eau-de-toilette-mixte/")
#makeLinkNotino()      
#data,size=makeNotino()
#test=bs4_scraping_notino()
#a=test.testctrat()

#Scrap force
"""e=0
t=3
while True:
    e+=1
    t+=50
    t=t%903
    try:
        makeIngredient(range(t-50,t))
    except:
        print("tenta n"+str(e))
"""
path =str(pathlib.Path(__file__).parent.absolute())+'\dataset.xlsx'
wb_obj = openpyxl.load_workbook(path.strip())
sheet_obj = wb_obj.active
for i in range(700):
    if pd.isnull(chget(i,data)["Marque"]):
        continue
    else:
        a=sheet_obj.cell(row=i+2,column=1)
        b=sheet_obj.cell(row=i+2,column=2)
        c=sheet_obj.cell(row=i+2,column=3)
        d=sheet_obj.cell(row=i+2,column=3)
        a.value=chget(i,data)["Marque"]
        b.value=chget(i,data)["Ingr"]
        c.value=chget(i,data)["Rating"]
        d.value=chget(i,data)["Reference"]
wb_obj.save('dataset.xlsx')

#wb_obj.save('data.xlsx')
    #if not pd.isnull(min_liste_ingr):
     #   for ii in min_liste_ingr:
      #      print(ii)
       #     liste_of_ingred=liste_of_ingred.append(ii)
        #print(min_liste_ingr)
#print(sum([1 for i in range(903) if pd.isnull(chget(i,data)["Ingr"]) or len(chget(i,data)["Ingr"].split(","))<5]))


